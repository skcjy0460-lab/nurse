import streamlit as st
import pandas as pd
from datetime import date
import calendar
import hashlib
import io

# ──────────────────────────────────────────────
# 페이지 기본 설정
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="일반병동 간호관리료 등급 산정",
    page_icon="🏥",
    layout="wide",
)

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 24px; font-weight: 800; color: #1a3a6b;
        border-bottom: 3px solid #1a3a6b; padding-bottom: 10px; margin-bottom: 16px;
        display: flex; align-items: baseline; gap: 12px;
    }
    .creator-badge { font-size: 13px; color: #888; font-weight: 500; }
    .section-title {
        font-size: 15px; font-weight: 700; color: #1a3a6b;
        background: #eef3fb; border-left: 5px solid #1a3a6b;
        padding: 7px 12px; margin: 14px 0 8px 0; border-radius: 0 6px 6px 0;
    }
    .result-card {
        background: #f0f7ff; border: 1.5px solid #aac8f0;
        border-radius: 10px; padding: 14px 20px; margin: 8px 0;
    }
    .grade-box {
        display: inline-block; font-size: 34px; font-weight: 900;
        padding: 12px 32px; border-radius: 12px; color: white; margin: 4px 0;
    }
    .grade-A { background: #5c1e91; }
    .grade-1 { background: #0d47a1; }
    .grade-2 { background: #1976d2; }
    .grade-3 { background: #2e7d32; }
    .grade-4 { background: #f57f17; }
    .grade-5 { background: #e65100; }
    .grade-6 { background: #b71c1c; }
    .kpi-label { font-size: 12px; color: #555; margin-bottom: 2px; }
    .kpi-value { font-size: 20px; font-weight: 700; color: #1a3a6b; }
    .kpi-unit  { font-size: 11px; color: #777; }
    .yellow-note {
        background: #fffde7; border: 1px solid #f9a825;
        border-radius: 6px; padding: 7px 12px; font-size: 12px; color: #5d4037;
        margin-bottom: 6px;
    }
    .footer {
        font-size: 13px; color: #555; text-align: center;
        margin-top: 30px; border-top: 1px solid #ddd; padding-top: 12px;
    }
    @media print {
        header, footer,
        [data-testid="stToolbar"], [data-testid="stSidebar"],
        [data-testid="stDecoration"], [data-testid="stStatusWidget"],
        .stButton > button, .stDownloadButton { display: none !important; }
        @page { size: A4 landscape; margin: 8mm; }
        html, body { margin: 0 !important; padding: 0 !important; }
        [data-testid="block-container"] {
            padding: 6px 10px !important; max-width: 100% !important; width: 100% !important;
        }
        * { font-size: 9px !important; line-height: 1.3 !important; }
        .main-title  { font-size: 13px !important; }
        .section-title { font-size: 10px !important; }
        .kpi-value { font-size: 12px !important; }
        .grade-box { font-size: 18px !important; padding: 6px 16px !important; }
        table, th, td { font-size: 9px !important; padding: 2px 4px !important; }
        .footer { margin-top: 10px !important; padding-top: 6px !important; }
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────────
QUARTER_RANGES = {
    "1분기 (12/15 ~ 3/14)": {"month_start": 12, "day_start": 15, "month_end": 3,  "day_end": 14},
    "2분기 (3/15 ~ 6/14)":  {"month_start": 3,  "day_start": 15, "month_end": 6,  "day_end": 14},
    "3분기 (6/15 ~ 9/14)":  {"month_start": 6,  "day_start": 15, "month_end": 9,  "day_end": 14},
    "4분기 (9/15 ~ 12/14)": {"month_start": 9,  "day_start": 15, "month_end": 12, "day_end": 14},
}

def get_quarter_dates(year, quarter_label):
    q = QUARTER_RANGES[quarter_label]
    if quarter_label.startswith("1"):
        start = date(year - 1, q["month_start"], q["day_start"])
        end   = date(year,     q["month_end"],   q["day_end"])
    else:
        start = date(year, q["month_start"], q["day_start"])
        end   = date(year, q["month_end"],   q["day_end"])
    return start, end, (end - start).days + 1

def calc_nurse_days(hire_date, status, q_start, q_end, resign_date=None):
    if status == "퇴사":
        if resign_date is None: return 0
        effective_end = min(resign_date, q_end)
    else:
        effective_end = q_end
    if hire_date > q_end: return 0
    if status == "퇴사" and resign_date is not None and resign_date < q_start: return 0
    effective_start = max(hire_date, q_start)
    return max(0, (effective_end - effective_start).days + 1)

def calc_parttime_weight(weekly_hours):
    if weekly_hours >= 40: return 1.0
    elif weekly_hours >= 36: return 0.8
    elif weekly_hours >= 32: return 0.6
    else: return 0.4

def determine_grade(ratio):
    pct = ratio * 100
    if   pct < 2.0: return "A등급"
    elif pct < 2.5: return "1등급"
    elif pct < 3.0: return "2등급"
    elif pct < 3.5: return "3등급"
    elif pct < 4.0: return "4등급"
    elif pct < 6.0: return "5등급"
    else:           return "6등급"

def grade_css(grade):
    return "grade-" + grade.replace("등급", "")

def month_label(base, offset):
    m = ((base.month - 1 + offset) % 12) + 1
    y = base.year + ((base.month - 1 + offset) // 12)
    return f"{y}년 {m}월"

# ──────────────────────────────────────────────
# 엑셀 파싱 함수
# ──────────────────────────────────────────────
def parse_excel_upload(file_bytes):
    import openpyxl
    from datetime import date as date_type, datetime as datetime_type

    def parse_date(v):
        if v is None: return None
        # datetime 먼저 (date의 서브클래스이므로 반드시 먼저)
        if isinstance(v, datetime_type): return v.date()
        if isinstance(v, date_type): return v
        s = str(v).strip()
        if not s: return None
        for fmt in [
            "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
            "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
        ]:
            try:
                from datetime import datetime as dt
                return dt.strptime(s, fmt).date()
            except: pass
        return None

    def to_int(v):
        if v is None: return None
        try: return int(float(str(v)))
        except: return None

    def to_str(v):
        return str(v).strip() if v is not None else None

    result = {"year": 2026, "quarter": None, "beds": 0,
              "patients": [0, 0, 0], "daytime": [], "night": []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        if "기본정보" in wb.sheetnames:
            ws = wb["기본정보"]
            y = to_int(ws["B5"].value);  result["year"]    = y if y else 2026
            q = to_str(ws["D5"].value);  result["quarter"] = q
            b = to_int(ws["F5"].value);  result["beds"]    = b if b is not None else 0
            for i, col in enumerate(["B","C","D"]):
                v = to_int(ws[f"{col}10"].value)
                result["patients"][i] = v if v is not None else 0

        if "주간간호사" in wb.sheetnames:
            ws2 = wb["주간간호사"]
            empty = 0
            for r in range(4, 104):
                hr = ws2.cell(r, 3).value
                if hr is None:
                    empty += 1
                    if empty >= 3: break
                    continue
                empty = 0
                hire   = parse_date(hr)
                resign = parse_date(ws2.cell(r, 4).value)
                status = to_str(ws2.cell(r, 5).value) or "근무"
                if status not in ["근무","퇴사"]: status = "근무"
                if hire:
                    result["daytime"].append({
                        "hire_date": hire,
                        "resign_date": resign if status == "퇴사" else None,
                        "status": status
                    })

        if "야간전담간호사" in wb.sheetnames:
            ws3 = wb["야간전담간호사"]
            empty = 0
            for r in range(4, 104):
                hr = ws3.cell(r, 3).value
                if hr is None:
                    empty += 1
                    if empty >= 3: break
                    continue
                empty = 0
                hire   = parse_date(hr)
                resign = parse_date(ws3.cell(r, 4).value)
                status = to_str(ws3.cell(r, 5).value) or "근무"
                if status not in ["근무","단시간근무","퇴사"]: status = "근무"
                hours  = to_int(ws3.cell(r, 6).value) or 40
                if hire:
                    result["night"].append({
                        "hire_date": hire,
                        "resign_date": resign if status == "퇴사" else None,
                        "status": status, "weekly_hours": hours
                    })
    except Exception as e:
        return None, str(e)
    return result, None

# ──────────────────────────────────────────────
# 세션 상태 초기화
# ──────────────────────────────────────────────
def init_session():
    if "year" not in st.session_state:       st.session_state["year"] = 2026
    if "quarter_idx" not in st.session_state: st.session_state["quarter_idx"] = 1
    if "beds" not in st.session_state:        st.session_state["beds"] = 0
    if "pat_0" not in st.session_state:       st.session_state["pat_0"] = 0
    if "pat_1" not in st.session_state:       st.session_state["pat_1"] = 0
    if "pat_2" not in st.session_state:       st.session_state["pat_2"] = 0
    if "daytime_nurses" not in st.session_state:
        st.session_state.daytime_nurses = [{"hire_date": None, "resign_date": None, "status": "근무"}]
    if "night_nurses" not in st.session_state:
        st.session_state.night_nurses = [{"hire_date": None, "resign_date": None, "status": "근무", "weekly_hours": 40}]

init_session()

# ──────────────────────────────────────────────
# 헤더
# ──────────────────────────────────────────────
st.markdown(
    '<div class="main-title">🏥 일반병동 간호관리료 등급 산정 시스템'
    '<span class="creator-badge">ㅣ 제작: 조정윤</span></div>',
    unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# 엑셀 업로드 — 세션 key에 직접 쓰기
# ──────────────────────────────────────────────
QUARTER_KEYS = list(QUARTER_RANGES.keys())

def apply_uploaded_data(parsed):
    st.session_state["year"] = parsed["year"]
    st.session_state["beds"] = parsed["beds"]
    for i, patient_count in enumerate(parsed["patients"]):
        st.session_state[f"pat_{i}"] = patient_count

    q = parsed["quarter"]
    quarter = q if q in QUARTER_KEYS else QUARTER_KEYS[1]
    st.session_state["quarter_idx"] = QUARTER_KEYS.index(quarter)
    st.session_state["quarter_sel"] = quarter

    widget_prefixes = (
        "d_hire_", "d_resign_", "d_status_",
        "n_hire_", "n_resign_", "n_status_", "n_hours_",
    )
    for key in list(st.session_state):
        if key.startswith(widget_prefixes):
            del st.session_state[key]

    daytime = parsed["daytime"] or [
        {"hire_date": None, "resign_date": None, "status": "근무"}
    ]
    night = parsed["night"] or [
        {"hire_date": None, "resign_date": None, "status": "근무", "weekly_hours": 40}
    ]
    st.session_state.daytime_nurses = daytime
    st.session_state.night_nurses = night

    for i, nurse in enumerate(daytime):
        st.session_state[f"d_hire_{i}"] = nurse["hire_date"]
        st.session_state[f"d_resign_{i}"] = nurse["resign_date"]
        st.session_state[f"d_status_{i}"] = nurse["status"]

    for i, nurse in enumerate(night):
        st.session_state[f"n_hire_{i}"] = nurse["hire_date"]
        st.session_state[f"n_resign_{i}"] = nurse["resign_date"]
        st.session_state[f"n_status_{i}"] = nurse["status"]
        st.session_state[f"n_hours_{i}"] = nurse["weekly_hours"]

with st.expander("📂 엑셀 파일로 데이터 자동 입력 (클릭하여 열기)", expanded=False):
    st.markdown(
        "<div style='background:#e3f2fd;border:1px solid #90caf9;border-radius:8px;"
        "padding:12px 16px;font-size:13px;'><b>사용 방법</b><br>"
        "1. 아래 버튼으로 <b>입력 양식 엑셀 파일</b>을 다운로드하세요.<br>"
        "2. 양식의 <span style='background:#FFF9C4;padding:1px 4px;border-radius:3px;'>"
        "노란색 셀</span>에 데이터를 입력 후 저장하세요.<br>"
        "3. 저장한 파일을 아래 업로드 칸에 올리면 모든 항목이 자동으로 채워집니다.</div>",
        unsafe_allow_html=True
    )
    try:
        with open("/mnt/user-data/outputs/간호관리료_데이터입력양식.xlsx", "rb") as tf:
            st.download_button(
                label="⬇️ 입력 양식 다운로드 (Excel)", data=tf.read(),
                file_name="간호관리료_데이터입력양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception:
        st.warning("양식 파일을 찾을 수 없습니다.")

    st.markdown("---")
    uploaded = st.file_uploader("작성한 엑셀 파일 업로드", type=["xlsx"], key="excel_upload")
    if uploaded is not None:
        uploaded_bytes = uploaded.getvalue()
        file_signature = hashlib.sha256(uploaded_bytes).hexdigest()
        reapply = st.button("업로드한 데이터 다시 적용", key="reapply_excel")
        if st.session_state.get("_applied_excel_signature") != file_signature or reapply:
            parsed, err = parse_excel_upload(uploaded_bytes)
            if err:
                st.error("파싱 오류: " + err)
            elif parsed:
                apply_uploaded_data(parsed)
                st.session_state["_applied_excel_signature"] = file_signature
                st.session_state["_excel_upload_message"] = (
                    "데이터 로드 완료! 주간 간호사 " + str(len(parsed["daytime"])) +
                    "명 / 야간전담 " + str(len(parsed["night"])) + "명 입력됨."
                )
                st.rerun()
        elif "_excel_upload_message" in st.session_state:
            st.success(st.session_state.pop("_excel_upload_message"))

# ──────────────────────────────────────────────
# ① 기본 정보
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">① 기본 정보</div>', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    year = st.number_input("연도", min_value=2020, max_value=2040, step=1, key="year")
with col2:
    quarter_label = st.selectbox("분기", QUARTER_KEYS,
                                 index=st.session_state["quarter_idx"], key="quarter_sel")
with col3:
    beds = st.number_input("운영 병상 수", min_value=0, max_value=500, step=1, key="beds")

q_start, q_end, q_days = get_quarter_dates(year, quarter_label)
st.info(f"📅 분기 기간: **{q_start}** ~ **{q_end}**  |  총 **{q_days}일**")

# ──────────────────────────────────────────────
# ② 월별 재원환자수
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">② 월별 재원환자수</div>', unsafe_allow_html=True)
st.markdown('<div class="yellow-note">🟡 <b>노란색 항목</b>: 일평균 재원환자 수는 자동 계산됩니다.</div>', unsafe_allow_html=True)

month_cols = st.columns(4)
total_patients = 0
for i in range(3):
    lbl = month_label(q_start, i)
    with month_cols[i]:
        st.markdown(f"**{lbl}**")
        pat = st.number_input("재원환자수", min_value=0, max_value=5000, step=1,
                              key=f"pat_{i}", label_visibility="collapsed")
        st.caption("월 재원환자수")
        total_patients += pat

with month_cols[3]:
    st.markdown("**일평균 재원환자 수** 🟡")
    avg_patients = total_patients / q_days if q_days > 0 else 0
    st.markdown(
        f'<div class="kpi-value">{avg_patients:.2f}</div>'
        f'<div class="kpi-unit">명/일 (자동계산)</div>', unsafe_allow_html=True
    )

# ──────────────────────────────────────────────
# ③ 주간 간호사
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">③ 주간(일반) 간호사 인력</div>', unsafe_allow_html=True)

c1, c2, _ = st.columns([1, 1, 6])
with c1:
    if st.button("➕ 주간 간호사 추가"):
        st.session_state.daytime_nurses.append({"hire_date": None, "resign_date": None, "status": "근무"})
with c2:
    if st.button("➖ 마지막 행 삭제") and len(st.session_state.daytime_nurses) > 1:
        st.session_state.daytime_nurses.pop()

hc = st.columns([0.4, 1.6, 1.6, 1.6, 1.6, 1.6])
hc[0].markdown("**#**"); hc[1].markdown("**입사일**"); hc[2].markdown("**퇴사일**")
hc[3].markdown("**상태**"); hc[4].markdown("**산정일수** 🟡"); hc[5].markdown("**환산인원** 🟡")

daytime_total = 0.0
for i, nurse in enumerate(st.session_state.daytime_nurses):
    cols = st.columns([0.4, 1.6, 1.6, 1.6, 1.6, 1.6])
    cols[0].markdown(f"{i+1}")

    hire = cols[1].date_input("입사일", value=nurse["hire_date"],
                              key=f"d_hire_{i}", label_visibility="collapsed")
    resign_disabled = (nurse["status"] != "퇴사")
    resign = cols[2].date_input("퇴사일", value=nurse["resign_date"],
                                key=f"d_resign_{i}", label_visibility="collapsed",
                                disabled=resign_disabled)
    status = cols[3].selectbox("상태", ["근무","퇴사"],
                               index=0 if nurse["status"] == "근무" else 1,
                               key=f"d_status_{i}", label_visibility="collapsed")

    resign_final = resign if status == "퇴사" else None
    if hire is not None:
        days_worked = calc_nurse_days(hire, status, q_start, q_end, resign_date=resign_final)
        weight = days_worked / q_days if q_days > 0 else 0.0
        cols[4].markdown(f'<div style="padding-top:8px;color:#1565c0;font-weight:600">{days_worked}일</div>', unsafe_allow_html=True)
        cols[5].markdown(f'<div style="padding-top:8px;color:#1565c0;font-weight:600">{weight:.2f}명</div>', unsafe_allow_html=True)
    else:
        weight = 0.0
        cols[4].markdown('<div style="padding-top:8px;color:#bbb">-</div>', unsafe_allow_html=True)
        cols[5].markdown('<div style="padding-top:8px;color:#bbb">-</div>', unsafe_allow_html=True)

    daytime_total += weight
    st.session_state.daytime_nurses[i] = {"hire_date": hire, "resign_date": resign_final, "status": status}

st.markdown(
    f'<div class="result-card">📊 <b>주간 간호사 3개월 평균 (환산합계)</b>: '
    f'<span class="kpi-value">{daytime_total:.2f}명</span></div>', unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# ④ 야간전담 간호사
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">④ 야간전담 간호사 인력</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🟡 단시간근무자 가중치 자동 적용: '
    '주40h↑=1.0 / 주36~40h=0.8 / 주32~36h=0.6 / 주32h↓=0.4</div>', unsafe_allow_html=True
)

c3, c4, _ = st.columns([1, 1, 6])
with c3:
    if st.button("➕ 야간 간호사 추가"):
        st.session_state.night_nurses.append({"hire_date": None, "resign_date": None, "status": "근무", "weekly_hours": 40})
with c4:
    if st.button("➖ 마지막 행 삭제 ", key="del_night") and len(st.session_state.night_nurses) > 1:
        st.session_state.night_nurses.pop()

nh = st.columns([0.4, 1.5, 1.5, 1.5, 1.5, 1.8, 1.5])
nh[0].markdown("**#**"); nh[1].markdown("**입사일**"); nh[2].markdown("**퇴사일**")
nh[3].markdown("**상태**"); nh[4].markdown("**근무시간**")
nh[5].markdown("**산정일수** 🟡"); nh[6].markdown("**환산인원** 🟡")

night_total = 0.0
for i, nurse in enumerate(st.session_state.night_nurses):
    cols = st.columns([0.4, 1.5, 1.5, 1.5, 1.5, 1.8, 1.5])
    cols[0].markdown(f"{i+1}")

    hire = cols[1].date_input("입사일", value=nurse["hire_date"],
                              key=f"n_hire_{i}", label_visibility="collapsed")
    resign_disabled = (nurse["status"] != "퇴사")
    resign = cols[2].date_input("퇴사일", value=nurse["resign_date"],
                                key=f"n_resign_{i}", label_visibility="collapsed",
                                disabled=resign_disabled)
    status_opts = ["근무","단시간근무","퇴사"]
    sidx = status_opts.index(nurse["status"]) if nurse["status"] in status_opts else 0
    status = cols[3].selectbox("상태", status_opts, index=sidx,
                               key=f"n_status_{i}", label_visibility="collapsed")
    weekly_h = cols[4].number_input("근무시간(h)", min_value=0, max_value=60,
                                    value=int(nurse.get("weekly_hours", 40)), step=1,
                                    key=f"n_hours_{i}", label_visibility="collapsed",
                                    disabled=(status != "단시간근무"))

    resign_final = resign if status == "퇴사" else None
    if hire is not None:
        days_worked = calc_nurse_days(hire, status, q_start, q_end, resign_date=resign_final)
        if status == "퇴사":
            weight = days_worked / q_days if q_days > 0 else 0.0
            eff_display = f"{days_worked}일"
        elif status == "단시간근무":
            pw = calc_parttime_weight(weekly_h)
            weight = (days_worked / q_days) * pw if q_days > 0 else 0.0
            eff_display = f"{days_worked}일 × {pw} = {days_worked*pw:.2f}일"
        else:
            weight = days_worked / q_days if q_days > 0 else 0.0
            eff_display = f"{days_worked}일"
        cols[5].markdown(f'<div style="padding-top:8px;color:#6a1b9a;font-weight:600;font-size:12px">{eff_display}</div>', unsafe_allow_html=True)
        cols[6].markdown(f'<div style="padding-top:8px;color:#6a1b9a;font-weight:600">{weight:.2f}명</div>', unsafe_allow_html=True)
    else:
        weight = 0.0
        cols[5].markdown('<div style="padding-top:8px;color:#bbb">-</div>', unsafe_allow_html=True)
        cols[6].markdown('<div style="padding-top:8px;color:#bbb">-</div>', unsafe_allow_html=True)

    night_total += weight
    st.session_state.night_nurses[i] = {"hire_date": hire, "resign_date": resign_final,
                                        "status": status, "weekly_hours": weekly_h}

st.markdown(
    f'<div class="result-card">📊 <b>야간전담 간호사 3개월 평균 (환산합계)</b>: '
    f'<span class="kpi-value" style="color:#6a1b9a">{night_total:.2f}명</span></div>',
    unsafe_allow_html=True
)

# ──────────────────────────────────────────────
# ⑤ 등급 산정 결과 보고서
# ──────────────────────────────────────────────
st.markdown('<div class="section-title">⑤ 등급 산정 결과 보고서</div>', unsafe_allow_html=True)
st.markdown('<div class="yellow-note">🟡 아래 항목은 모두 자동 계산됩니다.</div>', unsafe_allow_html=True)

total_nurses  = daytime_total + night_total
night_ratio   = (night_total / total_nurses * 100) if total_nurses > 0 else 0
patient_ratio = (avg_patients / total_nurses) if total_nurses > 0 else 0
grade         = determine_grade(patient_ratio / 100)

k1, k2, k3, k4, k5, k6 = st.columns(6)
def kpi(col, label, value, unit=""):
    col.markdown(
        f'<div class="result-card" style="text-align:center">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-unit">{unit}</div></div>', unsafe_allow_html=True
    )

kpi(k1, "🏥 운영 병상 수",         f"{beds}",                 "병상")
kpi(k2, "👥 일평균 재원환자 수",    f"{avg_patients:.2f}",     "명/일")
kpi(k3, "👩‍⚕️ 3개월 평균 간호사 수", f"{total_nurses:.2f}",     "명")
kpi(k4, "🌙 야간전담 간호사 수",    f"{night_total:.2f}",      "명")
kpi(k5, "📊 야간전담 간호사 비율",  f"{night_ratio:.2f}",      "%")
kpi(k6, "📐 환자대비 간호사수",     f"{patient_ratio:.2f}",    "(환자/간호사)")

st.markdown("---")
st.markdown(f"""
<div style="text-align:center; margin:14px 0;">
    <div style="font-size:16px; color:#555; margin-bottom:6px;">산정 등급</div>
    <span class="grade-box {grade_css(grade)}">{grade}</span>
    <div style="font-size:13px; color:#777; margin-top:8px;">
        환자대비 간호사수: <b>{patient_ratio:.2f} ({patient_ratio:.2f}%)</b>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("---")
st.markdown("#### 📋 등급 기준표")
grade_list = ["A등급","1등급","2등급","3등급","4등급","5등급","6등급"]
grade_table = pd.DataFrame({
    "등급": grade_list,
    "환자대비 간호사수 기준": [
        "2.0 미만","2.0 이상 ~ 2.5 미만","2.5 이상 ~ 3.0 미만",
        "3.0 이상 ~ 3.5 미만","3.5 이상 ~ 4.0 미만","4.0 이상 ~ 6.0 미만","6.0 이상"
    ],
    "현재": ["✅" if g == grade else "" for g in grade_list],
})
st.table(grade_table.set_index("등급"))

with st.expander("🔍 상세 계산 내역 보기"):
    st.markdown(f"""
| 항목 | 계산식 | 결과 |
|------|--------|------|
| 분기 일수 | {q_start} ~ {q_end} | **{q_days}일** |
| 총 재원환자수 | {total_patients}명 (3개월 합계) | |
| 일평균 재원환자수 | {total_patients} ÷ {q_days}일 | **{avg_patients:.2f}명** |
| 일반병동 간호사수 [주간+야간전담] | 각 간호사 근무일수 ÷ {q_days}일 합계 | **{daytime_total:.2f}명** |
| 야간전담 간호사 환산 | 근무일수 ÷ {q_days}일 × 가중치 합계 | **{night_total:.2f}명** |
| 3개월 평균 전체 간호사 수 | 일반병동 + 야간전담 | **{total_nurses:.2f}명** |
| 야간전담 간호사 비율 | {night_total:.2f} ÷ {total_nurses:.2f} × 100 | **{night_ratio:.2f}%** |
| 환자대비 간호사수 | {avg_patients:.2f} ÷ {total_nurses:.2f} | **{patient_ratio:.2f} ({patient_ratio:.2f}%)** |
| **산정 등급** | | **{grade}** |
""")

# ──────────────────────────────────────────────
# ⑥ AI 등급 진단 컨설팅 보고서
# ──────────────────────────────────────────────
import anthropic
import json

st.markdown('<div class="section-title">⑥ AI 등급 진단 및 컨설팅 보고서</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="yellow-note">🤖 현재 입력된 데이터를 바탕으로 AI가 등급 현황을 진단하고, '
    '등급 상향을 위한 구체적인 인력 충원 방안 및 경영 전략을 제안합니다.</div>',
    unsafe_allow_html=True
)

def next_grade_info(current_grade):
    order = ["A등급","1등급","2등급","3등급","4등급","5등급","6등급"]
    thresholds = {"A등급":(0,2.0),"1등급":(2.0,2.5),"2등급":(2.5,3.0),
                  "3등급":(3.0,3.5),"4등급":(3.5,4.0),"5등급":(4.0,6.0),"6등급":(6.0,999)}
    idx = order.index(current_grade)
    if idx == 0: return None, None, None
    prev = order[idx-1]
    _, upper = thresholds[prev]
    return prev, upper, thresholds[current_grade][0]

def nurses_needed_for_upgrade(avg_patients, total_nurses, current_grade):
    order = ["A등급","1등급","2등급","3등급","4등급","5등급","6등급"]
    uppers = [2.0, 2.5, 3.0, 3.5, 4.0, 6.0, 999]
    idx = order.index(current_grade)
    if idx == 0: return 0, "A등급"
    target = uppers[idx-1] - 0.01
    needed = avg_patients / target if target > 0 else 0
    return max(0, needed - total_nurses), order[idx-1]

next_g, next_upper, curr_lower = next_grade_info(grade)
add_nurses, upgrade_to = nurses_needed_for_upgrade(avg_patients, total_nurses, grade)

analysis_data = {
    "분기": quarter_label, "연도": year,
    "분기_시작일": str(q_start), "분기_종료일": str(q_end), "분기_일수": q_days,
    "운영_병상수": beds, "일평균_재원환자수": round(avg_patients, 2),
    "주간간호사_3개월평균": round(daytime_total, 2),
    "야간전담간호사_3개월평균": round(night_total, 2),
    "전체_간호사_3개월평균": round(total_nurses, 2),
    "야간전담_비율_퍼센트": round(night_ratio, 2),
    "환자대비_간호사수": round(patient_ratio, 2),
    "현재_등급": grade,
    "상위_목표등급": upgrade_to if upgrade_to else "현재 최고등급(A등급)",
    "등급_상향_추가필요_간호사수_환산": round(add_nurses, 2),
}

if st.button("🤖 AI 컨설팅 보고서 생성", type="primary", use_container_width=True):
    with st.spinner("AI가 데이터를 분석하고 컨설팅 보고서를 작성 중입니다..."):
        try:
            client = anthropic.Anthropic()
            system_prompt = """당신은 대한민국 의료기관 경영 전문 컨설턴트입니다.
특히 간호관리료 차등제 등급 관리에 특화된 전문가로서, '주식회사 메디엄'의 수석 컨설턴트입니다.
보고서는 전문적이고 수치 기반으로, 마크다운 형식으로 아래 구조를 따르세요:
# 일반병동 간호관리료 등급 진단 컨설팅 보고서
## 1. 현황 요약
## 2. 핵심 지표 분석
## 3. 등급 상향 전략 (단계별 인력 충원 시나리오)
### 시나리오 A: 주간 간호사 충원
### 시나리오 B: 야간전담 간호사 충원
### 시나리오 C: 혼합 충원 (최적안)
## 4. 재정적 효과 분석
## 5. 리스크 및 주의사항
## 6. 종합 권고사항 및 실행 로드맵"""

            user_prompt = (
                f"다음 데이터를 분석해 전문 컨설팅 보고서를 작성해주세요.\n\n"
                f"```json\n{json.dumps(analysis_data, ensure_ascii=False, indent=2)}\n```\n\n"
                f"현재 {grade}에서 {upgrade_to if upgrade_to else 'A등급'}으로 상향하기 위한 "
                f"구체적 간호사 충원 시나리오(주40h/주36h/주32h별)와 "
                f"야간전담 비율({night_ratio:.2f}%) 평가, 분기 내 입사 타이밍 전략을 포함해주세요."
            )

            report_placeholder = st.empty()
            full_report = ""
            with client.messages.stream(
                model="claude-sonnet-4-20250514", max_tokens=4000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}]
            ) as stream:
                for text in stream.text_stream:
                    full_report += text
                    report_placeholder.markdown(full_report)

            st.session_state["last_report"] = full_report
            st.success("✅ AI 컨설팅 보고서 생성 완료!")
        except Exception as e:
            st.error(f"AI 분석 오류: {str(e)}")
            st.info("💡 ANTHROPIC_API_KEY 환경변수를 확인해주세요.")

elif "last_report" in st.session_state:
    st.markdown(st.session_state["last_report"])
    st.info("💡 데이터를 변경한 후 버튼을 다시 누르면 새 보고서가 생성됩니다.")

st.markdown("---")
st.markdown("#### 📊 등급 상향 간호사 충원 시뮬레이션 (자동 계산)")

grade_order = ["A등급","1등급","2등급","3등급","4등급","5등급","6등급"]
uppers_map  = {"A등급":2.0,"1등급":2.5,"2등급":3.0,"3등급":3.5,"4등급":4.0,"5등급":6.0,"6등급":999}
curr_idx    = grade_order.index(grade)
sim_data    = []
for target_idx in range(0, curr_idx):
    tg = grade_order[target_idx]
    tr = uppers_map[tg] - 0.01
    needed     = avg_patients / tr if tr > 0 else 0
    additional = max(0, needed - total_nurses)
    sim_data.append({
        "목표 등급": tg,
        "필요 총 간호사(환산)": f"{needed:.2f}명",
        "추가 필요(환산)": f"{additional:.2f}명",
        "전일제 주간 충원 시": f"{additional:.2f}명 추가",
        "주36h 야간전담 충원 시": f"{additional/0.8:.2f}명 추가",
    })

if sim_data:
    st.dataframe(pd.DataFrame(sim_data), use_container_width=True, hide_index=True)
else:
    st.success("🎉 현재 최고 등급(A등급)입니다!")

# ──────────────────────────────────────────────
# 하단 푸터
# ──────────────────────────────────────────────
st.markdown(
    '<div class="footer">일반병동 간호관리료 등급 산정 시스템<br>'
    '<b>제작: 조정윤</b></div>', unsafe_allow_html=True
)
